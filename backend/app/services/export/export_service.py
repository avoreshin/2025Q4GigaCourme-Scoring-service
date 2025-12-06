from typing import Any
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO


class ExportService:
    """Service for exporting scoring reports to PDF and Excel"""
    
    def export_to_pdf(self, scoring: Any) -> bytes:
        """Export scoring report to PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30
        )
        story.append(Paragraph("Scoring Report", title_style))
        
        # Total Score
        score_style = ParagraphStyle(
            'Score',
            parent=styles['Normal'],
            fontSize=18,
            textColor=colors.HexColor('#0066cc'),
            spaceAfter=20
        )
        story.append(Paragraph(f"Total Score: {scoring.total_score}/100", score_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Breakdown
        story.append(Paragraph("Score Breakdown", styles['Heading2']))
        breakdown_data = [["Category", "Score"]]
        for category, score in scoring.breakdown.items():
            breakdown_data.append([category.replace("_", " ").title(), f"{score:.2f}"])
        
        breakdown_table = Table(breakdown_data)
        breakdown_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(breakdown_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Risks
        story.append(Paragraph("Top Risks", styles['Heading2']))
        for i, risk in enumerate(scoring.risks[:5], 1):
            risk_text = risk.get("description", str(risk))
            story.append(Paragraph(f"{i}. {risk_text}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Recommendations
        story.append(Paragraph("Recommendations", styles['Heading2']))
        for i, rec in enumerate(scoring.recommendations, 1):
            story.append(Paragraph(f"{i}. {rec}", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def export_to_excel(self, scoring: Any) -> bytes:
        """Export scoring report to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Scoring Report"
        
        # Header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        
        ws['A1'] = "Scoring Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        ws['A3'] = "Total Score"
        ws['B3'] = f"{scoring.total_score}/100"
        ws['B3'].font = Font(bold=True, size=14)
        
        # Breakdown
        ws['A5'] = "Category"
        ws['B5'] = "Score"
        for cell in ['A5', 'B5']:
            ws[cell].fill = header_fill
            ws[cell].font = header_font
            ws[cell].alignment = Alignment(horizontal='center')
        
        row = 6
        for category, score in scoring.breakdown.items():
            ws[f'A{row}'] = category.replace("_", " ").title()
            ws[f'B{row}'] = score
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            row += 1
        
        # Risks
        ws[f'A{row+1}'] = "Top Risks"
        ws[f'A{row+1}'].font = Font(bold=True, size=12)
        row += 2
        for i, risk in enumerate(scoring.risks[:5], 1):
            risk_text = risk.get("description", str(risk))
            ws[f'A{row}'] = f"{i}. {risk_text}"
            row += 1
        
        # Recommendations
        ws[f'A{row+1}'] = "Recommendations"
        ws[f'A{row+1}'].font = Font(bold=True, size=12)
        row += 2
        for i, rec in enumerate(scoring.recommendations, 1):
            ws[f'A{row}'] = f"{i}. {rec}"
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

